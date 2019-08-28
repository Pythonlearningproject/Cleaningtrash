import requests
from datetime import timedelta
import datetime
import pytz
import re
import openpyxl

headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/76.0.3809.100 Safari/537.36'}
r = requests.get('<Url>', auth=('<User>', '<Token>'), verify='/etc/ssl/certs/cacert.pem', headers=headers)
data = r.json()
for item in (data['jobs']):
    if item['color'] == 'red':
        api = (item['url'] + 'api/json?pretty=true')
        build_number = requests.get(api, auth=('<user>', '<token>'), verify='/etc/ssl/certs/cacert.pem', headers=headers)
        No = build_number.json()
        out = No['lastBuild']
        two = out['url']
        print "Pipeline: " + two
        api_1 = (two + 'api/json?pretty=true')
        output = requests.get(api_1, auth=('<user>', '<Token>'), verify='/etc/ssl/certs/cacert.pem', headers=headers)
        No_1 = output.json()
        d = No_1['timestamp'] / 1000
        date = datetime.datetime.utcfromtimestamp(int(d))
        timezone = pytz.timezone('Europe/Berlin')
        aware = timezone.localize(date)
        time = aware.strftime('%z')
        print "Job Started: ", aware + timedelta(hours=int(time[1:3]))
        print "Total Time Taken: ", "{:0>8}".format(str(datetime.timedelta(milliseconds=No_1['duration'])))
        url_2 = (two + '/consoleText')
        resp = requests.get(url_2, auth=('<User>', '<Token>'), verify='/etc/ssl/certs/cacert.pem', headers=headers)
        file = open("console_log.txt", "w")
        file.write(resp.text)
        file.close()
        infile = open('console_log.txt', "r")
        lines = infile.readlines()
        url_3 = (two + '/injectedEnvVars/api/json?pretty=true')
        resp_1 = requests.get(url_3, auth=('<User>', '<Token>'), verify='/etc/ssl/certs/cacert.pem', headers=headers)
        No_2 = resp_1.json()
        BN = No_2['envMap']
        print "Build Number: " + BN['BUILD_ID']
        logs = []
        for line in lines:
            if re.search(r"failure", line, re.IGNORECASE) or re.search(r"error", line, re.IGNORECASE) or re.search(r"failed", line, re.IGNORECASE) or re.search(r"not found", line, re.IGNORECASE):
                 logs.append(line)
        ex = openpyxl.load_workbook('test.xlsx')
        sh = ex.active
        sh['A2'] = BN['JOB_NAME']
        sh['B2'] = BN['BUILD_ID']
        sh['C2'] = aware + timedelta(hours=int(time[1:3]))
        sh['D2'] = "{:0>8}".format(str(datetime.timedelta(milliseconds=No_1['duration'])))
        hyper = two + '/console'
        sh.cell(row=2,column=5).value = '=HYPERLINK("%s","Logs")' % hyper
        sh.cell(row=2, column=6).value = "\n".join(logs)
        sh.insert_rows(2)
        ex.save('test.xlsx')
